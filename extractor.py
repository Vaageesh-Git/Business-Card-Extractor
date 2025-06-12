import os, base64, pytesseract, fitz, re, requests
import google.generativeai as genai
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from PIL import Image
from io import BytesIO
load_dotenv()

def extractor(url):
    response = requests.get(url)
    if response.status_code != 200:
        raise Exception("Failed to download the file from URL")
    content_type = response.headers.get("Content-Type")
    file_type = content_type.split(";")[0] if content_type else None

    API_KEY = os.getenv("GEMINI_API_KEY")
    genai.configure(api_key=API_KEY)
    
    column_heading = ["Company", "Phone", "Email", "Address", "Name","Designation", "Website"]
    
    print("File Type", file_type) 
    extracted_data = []

    pages = []
    if file_type == "application/pdf":
        doc = fitz.open("pdf", response.content)
        print(f"📄 Extracting {len(doc)} pages from PDF...")
        
        for page_number in range(len(doc)):
            page = doc.load_page(page_number)
            pix = page.get_pixmap()
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            pages.append(img)

        file_type = "image/png"

    elif file_type in ["image/jpeg", "image/png"]:
        try:
            img = Image.open(BytesIO(response.content))
            pages = [img]
        except Exception as e:
            print("❌ Error opening image:", e)
            return
    else:
        print("⚠️ Unsupported file format:", file_type)
        return

    for img in pages:
        data = {column: "" for column in column_heading}
        img_bytes = BytesIO()
        img = img.convert("RGB")
        img.save(img_bytes, format="PNG")
        img_bytes = img_bytes.getvalue()
        base_64_data = base64.b64encode(img_bytes)
        base_64_str = base_64_data.decode("utf-8")

        try:
            model = genai.GenerativeModel("models/gemini-2.0-flash")
            response = model.generate_content([
                '''You are an expert at reading business cards.
                From the image, extract and return only the contact information in thisstructured format:
                Company: [Company Name]
                Phone: [Phone Number] 
                Email: [Email Address]
                Address: [Full Address]
                Name: [Full Name]
                Designation: [Designation]
                Website: [Website URL] 
                If the language of the card is differnt from English, convert the details into English language and then give the details
                Return plain text only. Do not use asterisks, bullets, or markdown formatting.
                Do not include any explanations or extra text. If any field is not found, leaveit blank.''',
                {
                    "mime_type": file_type,
                    "data": base_64_str
                }
            ])
            extracted_text = response.text
            print("\n✅ Extracted Text:\n", extracted_text)
        
            for line in extracted_text.strip().split('\n'):
                if ":" in line:
                    split_lst = line.split(":",1)
                    data[split_lst[0].strip()] = split_lst[1].strip()
            print(data)

        except Exception as e:
            print("❌ Error:", e)

            try:
                raw_text = pytesseract.image_to_string(img)
                print("Raw OCR Text :", raw_text)

                for line in raw_text.split("\n"):
                    line = line.strip()

                    if not data["Email"]:
                        match = re.search(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", line)
                        if match:
                            data["Email"] = match.group()
                            continue

                    if not data["Website"]:
                        match = re.search(r"(https?://)?(www\.)?[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", line)
                        if match:
                            website = match.group()
                            if not website.startswith("http"):
                                website = "http://" + website
                            data["Website"] = website
                            continue

                    if not data["Phone"]:
                        match = re.search(r"(?:(?:\+|00)\d{1,3})?\s?(?:\(?\d{2,4}\)?[\s-]?)?\d{5,}", line)
                        if match:
                            data["Phone"] = match.group().strip()
                            continue

                    if not data["Company"]:
                        if any(kw in line for kw in ["Pvt", "Ltd", "LLC", "Inc", "Solutions", "Technologies", "Corporation"]):
                            data["Company"] = line
                            continue
                        elif line.isupper() and len(line.split()) <= 4 and len(line) > 3:
                            data["Company"] = line
                            continue

            except Exception as e:
                print("OCR Error", e)
        
        extracted_data.append(data)
        excel_file = "extracted_data.xlsx"
        if os.path.exists(excel_file):
            wb = load_workbook(excel_file)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Business Card"
            ws.append(column_heading)
            
        row_data = []
        for column in column_heading:
            d = data.get(column, "")
            row_data.append(d)
        print(row_data)
        if any(dt != "" for dt in row_data):
            ws.append(row_data)
        wb.save(excel_file)
extractor("")          # add the url here inside the string 

